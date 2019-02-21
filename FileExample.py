from openpyxl import *
#print the data 
filepath="C:\PythonProjects\ListExample\Book1.xlsx"

wb=load_workbook(filepath)
sheet=wb.active

max_row=sheet.max_row

max_col=sheet.max_column

file="C:\PythonProjects\ListExample\TextFile.txt"

'''
for i in range(1,max_row+1):
    for j in range(1,max_col+1):
        print(sheet.cell(row=i,column=j).value)
#read the data from txt file and write to excel file
'''

list1=[]
def txtfile():

    file=open("C:\PythonProjects\ListExample\TextFile.txt")
    for i in file.readlines():
        list1.append(i.split())
    print(list1)
    return list1

def writtoexl(mydata):

    filepath="C:\PythonProjects\ListExample\Book1.xlsx"
    wb=load_workbook(filepath)
    sheet=wb.active

    for r in range(1,len(mydata)+1):
        for c in range(1,len(mydata[0])+1):
            sheet.cell(row=r, column=c).value =  mydata[r-1][c-1]
    wb.save(filepath)
    data=mydata
    print(data)



mydata = txtfile()
writtoexl(mydata)
'''
#two decorator for single function

def Fun1(func):
    def inner1():
        print("in decoratoor in 1")
        return "deco1"+func()

    return inner1

def fun2(func):
    def inner2():
        print("inside the decoraotr 2")
        return "deco2" + func()

    return inner2
@Fun1
@fun2
def greet():
    print("inside greet")
    return "greet"
print(greet())
'''












