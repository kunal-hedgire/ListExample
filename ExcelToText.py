#excel to file and db

"https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f"
from  openpyxl import *
import pymysql

conn= pymysql.connect(host='localhost',port=3306,user='root',passwd='tiger',db='basic')
print("Connect")

cur=conn.cursor()

filepath="C:\PythonProjects\ListExample\Book1.xlsx"



wb=load_workbook(filepath)
sheet=wb.active
max_row = sheet.max_row
# get max column count
max_column = sheet.max_column
# iterate over all cells
# iterate over all rows
outfile=open("C:\PythonProjects\ListExample\TextFile.txt","w")
#if sheet.cell(row=2,col=4).value:



for i in range(2,max_row+1):
    #print(sheet.cell(row=i,column=3).value)
    #print(sheet.cell(row=i,column=4).value)
    if sheet.cell(row=i,column=3).value == sheet.cell(row=i,column=4).value:
        print("equal")
        #print(query)
        query = "insert into stud (firstname,course,paid) values('" + sheet.cell(row=i,column=1).value + "','" + sheet.cell(row=i, column=2).value + "','yes')"
        #values = (firstname, course)
        cur.execute(query)
    else:
        print("not equal")
        outfile.writelines("name = " + sheet.cell(row=i, column=1).value + "-- course = " + sheet.cell(row=i,column=2).value + "Fees Rem= " + str(sheet.cell(row=i, column=3).value - sheet.cell(row=i, column=4).value)+"\n" )
            #cur.close()


conn.commit()
conn.close()



'''
#Excel data to txt
from openpyxl import *

filepath="C:\PythonProjects\ListExample\Book1.xlsx"

wb=load_workbook(filepath)
sheet=wb.active

max_row = sheet.max_row
# get max column count
max_column = sheet.max_column
#print(max_column)
outfile=open("C:\PythonProjects\ListExample\TextFile.txt","w")

for i in range(2,max_row+1):
    #pass
    #for j in range(1,max_column+1):
        #print(sheet.cell(row=i, column=j).value)
    outfile.writelines("name = " + sheet.cell(row=i, column=1).value + "-- course = " + sheet.cell(row=i,column=2).value + "Fees = " + str(sheet.cell(row=i, column=3).value)+"\n")
'''


