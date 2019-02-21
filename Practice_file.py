from openpyxl import *

filepath="C:\PythonProjects\ListExample\Book1.xlsx"

wb=load_workbook(filepath)
sheet=wb.active
outfile=open("C:\PythonProjects\ListExample\\New Text Document.txt","w")
for i in range(2,sheet.max_row+1):
    outfile.writelines("name:---"+sheet.cell(row=i,column=1).value+"-====course==="+sheet.cell(row=i,column=2).value+"======fees:==="+str(sheet.cell(row=i,column=3).value)+"\n")
    #for j in range(1,sheet.max_column+1):
        #print(sheet.cell(row=i,column=j).value)


